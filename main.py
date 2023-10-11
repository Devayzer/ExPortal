import re
import csv
from datetime import datetime
import argparse
import openpyxl
import chardet

def detect_encoding(file_path):
    with open(file_path, 'rb') as file:
        result = chardet.detect(file.read())
    return result['encoding']

def parse_txt_file(txt_file, encoding):
    with open(txt_file, "r", encoding=encoding) as file:
        txt_data = file.read()

    # Визначаємо регулярні вирази для вилучення URL, Title та Visited On
    url_pattern = r"URL\s*:\s*(.*?)\n"
    title_pattern = r"Title\s*:\s*(.*?)\n"
    visited_pattern = r"Visited On\s*:\s*(.*?)\n"

    # Використовуємо регулярні вирази для пошуку даних
    urls = re.findall(url_pattern, txt_data)
    titles = re.findall(title_pattern, txt_data)
    visited_dates = re.findall(visited_pattern, txt_data)

    # Перетворюємо дати у зручний формат
    formatted_dates = [datetime.strptime(date, "%d.%m.%Y %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S") for date in visited_dates]

    return urls, titles, formatted_dates

def main():
    parser = argparse.ArgumentParser(description="Process a TXT file with various data and convert it to XLSX.")
    parser.add_argument("input_file", help="Path to the input TXT file")
    parser.add_argument("output_file", help="Path to the output XLSX file")

    args = parser.parse_args()

    # Визначаємо кодировку і читаємо файл з нею
    encoding = detect_encoding(args.input_file)
    urls, titles, visited_dates = parse_txt_file(args.input_file, encoding)

    # Створюємо новий Excel-файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Встановлюємо ширину стовпців
    sheet.column_dimensions["A"].width = 90
    sheet.column_dimensions["B"].width = 60
    sheet.column_dimensions["C"].width = 30

    # Створюємо стиль шрифту з розміром тексту 16px та жирним шрифтом для заголовків
    header_font = openpyxl.styles.Font(size=16, bold=True)

    # Створюємо стиль шрифту з розміром тексту 14px для інших даних
    data_font = openpyxl.styles.Font(size=14)

    # Створюємо стиль для заливки кольором #C9C9C9
    header_fill = openpyxl.styles.PatternFill(start_color="C9C9C9", end_color="C9C9C9", fill_type="solid")

    # Записуємо дані в аркуш Excel з правильними назвами стовпців та стилями
    sheet.append(["Адреса", "Назва сторінки", "Дата та час"])

    for url, title, visited_date in zip(urls, titles, visited_dates):
        sheet.append([url, title, visited_date])

    for cell in sheet["1"]:
        cell.font = header_font  # Встановлюємо розмір тексту 16px та жирний стиль для першого рядка
        cell.fill = header_fill  # Встановлюємо заливку кольором #C9C9C9

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal="center")
            cell.font = data_font  # Встановлюємо розмір тексту 14px для інших даних

    # Встановлюємо вирівнювання тексту по центру для всіх стовпців
    for column in sheet.columns:
        for cell in column:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Додаємо межі для всієї таблиці
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                   right=openpyxl.styles.Side(style="thin"),
                                   top=openpyxl.styles.Side(style="thin"),
                                   bottom=openpyxl.styles.Side(style="thin"))
    
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=3):
        for cell in row:
            cell.border = border

    # Переіменовуємо аркуш
    sheet.title = args.input_file.split(".")[0]

    # Зберігаємо Excel-файл
    workbook.save(args.output_file)

    print(f"Дані з файлу {args.input_file} були успішно записані у файл {args.output_file} з аркушем '{sheet.title}'")

if __name__ == "__main__":
    main()